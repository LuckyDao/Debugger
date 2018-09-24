
#TODO:  180617_10 LG make browser options choosable incl chrome ff and ie
#TODO:  180618_17 LG Should handle SSL certificates properly at some point - will need to find my certificates on the computer - requests package looks like it can handle this well
    #https://stackoverflow.com/questions/18061640/ignore-certificate-validation-with-urllib3  Could switch between HTTPSConnectionPool and PoolManager
    #HACK:  https://stackoverflow.com/questions/27981545/suppress-insecurerequestwarning-unverified-https-request-is-being-made-in-pytho
#TODO:  180623_20 MD Rather than try except, for non critical errors, I could log them
    #updating s_code check - if it finds utag.js, have to go into the utag code itself?  Not sure
#TODO:  180624_24 VLG Should I use the htmlUnit driver rather than FF or Chrome default? https://www.seleniumhq.org/docs/03_webdriver.jsp#selenium-webdriver-s-
#TODO:  180630_28 MD Sometimes I receive a 504 or 304 error, and my URL if returns a "STATUS OK"  this is incorrect --503 IS FOR CHECKOUT PAGES - SHOULD KEEP THESE
#TODO:  180630_29 SM print my evars/props, but also pass them into arrays
#TODO:  180630_30 SM Randomize the sleep amount, but use implicit wait (todo 21)
    #wait http://selenium-python.readthedocs.io/waits.html
#TODO: 180705_32 LG check if an instance of the browser is already open, if so, don't open a new one
#TODO:  180705_33 MD LANDSEND DOESN"T WORK - it has a popup modal - must kill the popup modal --STILL AN ISSUE??? CHECK
#TODO:  180730_37 SM change globals.* to glb.*
#TODO:  180730_38 MD MOBILE UA ISN'T WORKING?

#TODO:  180924_39 URL Test No longer works
#TODO:  180924_40 Kessel Run has an else that may not be required

print("**ACTUAL CODE...Last Commit:  9/24/18 5:51PM")

#TODO CURRENT PROJECT:  180808_40 Create and connect mechanism to output to Excel:
    # Create an excel spreadsheet
    #write page 1 to a sheet, write page 2 to a sheet, etc.
    #toggle for output to one sheet or output to many sheets?
    # Show Output Path and file name in an entry box
    # how to work with excel https://www.datacamp.com/community/tutorials/python-excel-tutorial
    #openpyexcel https://openpyxl.readthedocs.io/en/stable/tutorial.html
    #TODO -- I left off Issues with setting active sheet in openpyxl



from selenium import webdriver
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.firefox.options import Options as Options #allows for some browser options
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities #allows for some browser options
from time import sleep #allows for waiting w/ promise
import requests #used for URL functions and validity
import urllib3 #used to kill Insecure https request warnings- should solve properly for this

import openpyxl #https://openpyxl.readthedocs.io/en/stable/tutorial.html

#these imports and two lines may be necessary to import globals.
# import sys
# import os
#
# file_dir = os.path.dirname("C:/Python36/Env/PacketSniffer/")
# sys.path.append(file_dir)

import globals
print("...Import complete...")


#9/17/18 edit - try without headless
#CREATE OPTIONS AND CAPABILITIES OBJECTS
ops = Options()
caps = DesiredCapabilities.FIREFOX.copy()
profile = webdriver.FirefoxProfile()
prof = webdriver.FirefoxProfile()

#MODIFY OPTIONS AND CAPABILITIES TO BE HEADLESS AND EAGER LOAD STRATEGY, ATTEMPTING TO ADD IN UA
#removed Ops and UA for debugging 9/17/2018
caps.update({'pageLoadStrategy':'eager'})
ops.headless=True

#only works in scratch file
#prof.set_preference("general.useragent.override", "Android 4.4; Mobile;")

#required geckodriver v20???  May end timeout issues
browser = webdriver.Firefox(executable_path="C:\\Python36\\Env\\PacketSniffer\\Drivers for Selenium Library\\geckodriver20.exe", capabilities=caps, timeout=10, options = ops)
print("...Globals are set...")

# create an array of the vars and props for export to excel
#will need to take each page and put it on a separate excel sheet
MagicArray = []

#Sets up the workbook for output
Wb = openpyxl.Workbook()


def WritePageCode():

    if globals.OutputPath == "":
        print("Output Path not set!!!!")
    else:
        #Wb.save(str(globals.OutputPath) + "\Testfile.xlsx") ##Can put this at the end of the Fx
        sheetcount =0
        Wb.create_sheet()
        Wb.remove_sheet("Sheet1")

    #THESE WILL BE CALLED BY KESSELRUN

        Ws = Wb.create_sheet("homepage")
        cellrange = Ws["A1"]
        cellrange.value = "Hello World"


def URLTest():
    '''TESTS THE URL FOR STANDARD HTTP ERRORS, DISABLES SSL CHECKS URL CHECK ISN'T CURRENTLY WORKING'''
    try:
        #hard disable SSL warning
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        r = requests.head(globals.TargetURL, verify=False)
        #URL Status ok?
        if r.status_code == 200 or 301 or 302:
            print("\nURL STATUS OK:  "+str(r.status_code))
        else:
            print("URL STATUS NOT 200:  " + str(r.status_code))
            return
    except requests.exceptions.RequestException as e:
        print("GET ENCOUNTERED A FAILURE"+str(e))

def StandardCodeChecks():
    '''TESTS FOR EXISTENCE OF ADOBE COOKIE AND S_CODE, ELSE QUITS'''
    CookieExists = browser.get_cookie(globals.OrgID)
    MCIDExists = browser.execute_script("return("+globals.AdobeNamespace+"marketingCloudVisitorID"+");")

    #check for cookie, if not, nothing else matters  Should build if cookie changes page to page as well
    if CookieExists is not None:
        # InitialVisitorCookie = browser.get_cookie(globals.OrgID)
        # #FUTURE STATE:  if CurrentCookie == NULL Then Continue else if Initial Visitor Cookie != CurrentCookie throw error
        print("...Cookie Was Found...")
        sleep(2)
        if MCIDExists is None or MCIDExists=="":
            print("MCID NOT SET!")
        else:
            print("...MCID Was Set...")
    else:
        print("COOKIE OR ORG ID NOT FOUND!")


    # check for presence of s_code.js, (this only works without data layer or DTM... update IT later)
    if len(browser.find_elements_by_xpath("//script[contains(@src, 's_code.js')]"))>0:
        print("...s_code exists...")
    else:
        print("s_code does not exist!...")
        return

def ScriptExecution(SVarStr,PrintMe=True):
    '''GRABS THE TAG VALUES, APPENDS THEM TO AN ARRAY.  INTENDED TO OPERATE ON EACH TAG.  TAKES TAG NAME AS PARAM 1 (pageName) for example'''
    #TODO:  IF PRINTME IS TRUE, PRINT TO CONSOLE AS WELL AS ARRAY
    SVarVal = browser.execute_script("return ("+globals.AdobeNamespace+SVarStr+");")
    if SVarVal is None:
        x=0
    elif SVarVal == "" and PrintMe==True:
        MagicArray.append(str(SVarStr)+":"+"  ---")
        print(str(SVarStr)+":"+"  ---")

    elif PrintMe==True:
        MagicArray.append(str(SVarStr)+":"+"  "+str(SVarVal))
        print(str(SVarStr)+":"+"  "+str(SVarVal))

def PullPage():
    '''GRABS THE PAGE, RUNS SCRIPTEXECUTION FUNCTION ON EACH SET OF TAGS, CHECKS EACH PAGE FOR STANDARD CODE ELEMENTS'''
    browser.get(globals.TargetURL)
    print("...Getting URL...")
    sleep(4)

    #RUN STANDARD CHECKS FOR ADOBE CODE
    StandardCodeChecks()

#DO NOT DELETE - THIS CODE WORKS
    try:
        print("**************")
        ScriptExecution("pageName")
        ScriptExecution("pageURL")
        print("**************")
        ScriptExecution("marketingCloudVisitorID")
        ScriptExecution("products")
        for i in range(0, 25, 1):
            ScriptExecution("prop"+str(i))
            ScriptExecution("eVar" + str(i))

    except WebDriverException as e:
        print("WebDriver Exception on {}".format(globals.TargetURL))
        print("\n"+str(e))

def KesselRun():
    '''LOOPS THROUGH THE URLARRAY, EXECUTES URLTEST, PULLDATA (w/ Script Execution) and then WritePageCode IN SEQUENCE, KNOWS WHEN THE URL CHANGES'''
    for m in range(0, len(globals.URLArray)):
        globals.TargetURL = globals.URLArray[m]

        #disabled 9/17/18 - may be blocking code
        #URLTest()

        PullPage()
        #WritePageCode()

        m=m + 1
    else:
        #THIS BIT MAY NOT BE NEEDED
        globals.URLArray = [""]
        globals.OrgID=""
        print("\n12 Parsecs!")


print("End of ScrapeScript")
